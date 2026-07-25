from __future__ import annotations

import csv
import io
import re
import unicodedata
import uuid
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_datetime

from apps.audit.services import log_audit_event
from apps.sales.models import EcommerceSale, FeriaSale, KommoSale, ManualSale, SaleSource
from apps.sales.services.normalization import promote_to_consolidated
from apps.sales.services.payment_methods import resolve_payment_method
from apps.sales.services.fulfillment import apply_fulfillment, normalize_fulfillment_type
from apps.sales.kit_types import normalize_kit_type

CANONICAL_FIELDS = [
    "external_id",
    "source",
    "customer_name",
    "email",
    "phone",
    "id_number",
    "address_raw",
    "city_raw",
    "state_raw",
    "total_value",
    "amount_shipping",
    "payment_account",
    "commercial_raw",
    "qty_dorados",
    "qty_plateados",
    "tipo_dorados",
    "tipo_plateados",
    "status",
    "closed_at",
    "order_notes",
    "fulfillment_type",
    # Históricos ya despachados
    "tracking_number",
    "tracking_url",
    "shipping_cost",
    "label_url",
    "carrier",
    "sent_at",
    "shipment_status",
]

HEADER_ALIASES: dict[str, str] = {
    "id": "external_id",
    "external_id": "external_id",
    "order_id": "external_id",
    "lead_id": "external_id",
    "deal_id": "external_id",
    "formateador_de_id": "external_id",
    "canal": "source",
    "source": "source",
    "income_source": "source",
    "fuente": "source",
    "fuente_de_ingreso": "source",
    "cliente": "customer_name",
    "customer_name": "customer_name",
    "nombre": "customer_name",
    "nombre_del_negocio": "customer_name",
    "deal_name": "customer_name",
    "email": "email",
    "correo": "email",
    "correo_asociado": "email",
    "associated_contact": "email",
    "telefono": "phone",
    "phone": "phone",
    "celular": "phone",
    "cedula": "id_number",
    "cc": "id_number",
    "id_number": "id_number",
    "direccion": "address_raw",
    "address": "address_raw",
    "address_raw": "address_raw",
    "ciudad": "city_raw",
    "city": "city_raw",
    "city_raw": "city_raw",
    "departamento": "state_raw",
    "state_raw": "state_raw",
    "valor": "total_value",
    "total": "total_value",
    "total_value": "total_value",
    "transporte": "amount_shipping",
    "trasporte": "amount_shipping",  # typo en export histórico
    "shipping": "amount_shipping",
    "amount_shipping": "amount_shipping",
    "cuenta": "payment_account",
    "cuenta_bancaria": "payment_account",
    "payment_account": "payment_account",
    "vendedor": "commercial_raw",
    "comercial": "commercial_raw",
    "commercial_raw": "commercial_raw",
    "dorados": "qty_dorados",
    "cantidad_dorados": "qty_dorados",
    "qty_dorados": "qty_dorados",
    "plateados": "qty_plateados",
    "cantidad_plateados": "qty_plateados",
    "qty_plateados": "qty_plateados",
    "tipo_dorados": "tipo_dorados",
    "tipo dorados": "tipo_dorados",
    "tipo_plateados": "tipo_plateados",
    "tipo plateados": "tipo_plateados",
    "status": "status",
    "estado": "status",
    "status_de_la_orden": "status",
    "fecha": "closed_at",
    "closed_at": "closed_at",
    "fecha_cierre": "closed_at",
    "fecha_de_cierre": "closed_at",
    "notas": "order_notes",
    "order_notes": "order_notes",
    "fulfillment_type": "fulfillment_type",
    "tipo_entrega": "fulfillment_type",
    "entrega": "fulfillment_type",
    "fulfillment": "fulfillment_type",
    "guia": "tracking_number",
    "tracking": "tracking_number",
    "tracking_number": "tracking_number",
    "numero_guia": "tracking_number",
    "numero_de_guia": "tracking_number",
    "nro_guia": "tracking_number",
    "tracking_url": "tracking_url",
    "url_seguimiento": "tracking_url",
    "link_seguimiento": "tracking_url",
    "costo_guia": "shipping_cost",
    "shipping_cost": "shipping_cost",
    "costo_envio": "shipping_cost",
    "label_url": "label_url",
    "etiqueta": "label_url",
    "pdf_guia": "label_url",
    "carrier": "carrier",
    "transportadora": "carrier",
    "sent_at": "sent_at",
    "fecha_envio": "sent_at",
    "fecha_enviado": "sent_at",
    "fecha_de_generacion_de_guia": "sent_at",
    "shipment_status": "shipment_status",
    "estado_envio": "shipment_status",
    "estado_guia": "shipment_status",
    "enviado": "shipment_status",
}

SOURCE_MODELS = {
    SaleSource.ECOMMERCE: EcommerceSale,
    SaleSource.KOMMO: KommoSale,
    SaleSource.FERIAS: FeriaSale,
    SaleSource.MANUAL: ManualSale,
}


def _norm_header(h: str) -> str:
    """Lowercase, spaces→_, strip accents (NÚMERO DE GUÍA → numero_de_guia)."""
    s = (h or "").strip().lower().replace(" ", "_")
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return s


def detect_mapping(headers: list[str]) -> dict[str, str | None]:
    """Map canonical field -> CSV header name (or None)."""
    mapping: dict[str, str | None] = {f: None for f in CANONICAL_FIELDS}
    for header in headers:
        key = HEADER_ALIASES.get(_norm_header(header))
        if key and mapping.get(key) is None:
            mapping[key] = header
    return mapping


def _scalar_str(val: Any) -> str:
    """Excel-safe stringify: 76116174690.0 → '76116174690', bools/dates clean."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "true" if val else "false"
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val)
    if isinstance(val, datetime):
        return val.replace(microsecond=0).isoformat(sep=" ")
    if isinstance(val, date):
        return val.isoformat()
    text = str(val).strip()
    if re.fullmatch(r"-?\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def _customer_display_name(raw: str) -> str:
    """'Ecommerce | 4435 | Maribel Henao' → 'Maribel Henao'."""
    name = (raw or "").strip()
    if "|" in name:
        parts = [p.strip() for p in name.split("|") if p.strip()]
        if len(parts) >= 2:
            return parts[-1]
    return name


def _clip(value: Any, max_len: int) -> str:
    text = _scalar_str(value)
    if max_len > 0 and len(text) > max_len:
        return text[:max_len]
    return text


def _sanitize_id_number(raw: str) -> str:
    """Drop Woo JSON / garbage that sometimes lands in the CC column."""
    text = (raw or "").strip()
    if not text:
        return ""
    if text.startswith("{") or text.startswith("["):
        return ""
    if len(text) > 64:
        return ""
    # Keep typical document tokens; otherwise clip.
    return text[:64]


def _sanitize_phone(raw: str) -> str:
    text = (raw or "").strip()
    return text[:64]


def _dec(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    raw = str(value).strip().replace("$", "").replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return Decimal(raw)
    except InvalidOperation:
        raise ValueError(f"Monto inválido: {value}")


def _int(value: Any) -> int:
    if value is None or value == "":
        return 0
    return int(float(str(value).strip().replace(",", ".")))


def _parse_closed_at(value: Any):
    if not value:
        return timezone.now()
    if isinstance(value, datetime):
        dt = value
        if timezone.is_naive(dt):
            return timezone.make_aware(dt, timezone.get_current_timezone())
        return dt
    if isinstance(value, date) and not isinstance(value, datetime):
        return timezone.make_aware(datetime.combine(value, time(12, 0)), timezone.get_current_timezone())

    text = str(value).strip()
    isoish = text.replace(" ", "T", 1) if " " in text and "T" not in text else text
    dt = parse_datetime(isoish)
    if dt:
        if timezone.is_naive(dt):
            return timezone.make_aware(dt, timezone.get_current_timezone())
        return dt
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y",
        "%m/%d/%Y",
    ):
        try:
            naive = datetime.strptime(text[:26], fmt)
            return timezone.make_aware(naive, timezone.get_current_timezone())
        except ValueError:
            continue
    try:
        d = date.fromisoformat(text[:10])
        return timezone.make_aware(datetime.combine(d, time(12, 0)), timezone.get_current_timezone())
    except ValueError:
        return timezone.now()


def _normalize_source(raw: str | None) -> str:
    s = (raw or "MANUAL").strip().upper()
    aliases = {
        "E-COMMERCE": SaleSource.ECOMMERCE,
        "ECOMMERCE": SaleSource.ECOMMERCE,
        "WOO": SaleSource.ECOMMERCE,
        "WOOCOMMERCE": SaleSource.ECOMMERCE,
        "KOMMO": SaleSource.KOMMO,
        "FERIA": SaleSource.FERIAS,
        "FERIAS": SaleSource.FERIAS,
        "MANUAL": SaleSource.MANUAL,
        "MANUALES": SaleSource.MANUAL,
    }
    return aliases.get(s, SaleSource.MANUAL)


def parse_csv_text(text: str) -> tuple[list[str], list[dict[str, str]]]:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = reader.fieldnames or []
    rows = [dict(row) for row in reader]
    return headers, rows


def rows_to_csv_text(headers: list[str], rows: list[dict[str, Any]]) -> str:
    """Serialize header+row dicts to CSV text for reuse of dry_run/commit_csv."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({h: "" if row.get(h) is None else str(row.get(h)) for h in headers})
    return buf.getvalue()


def parse_xlsx_bytes(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Read first sheet of an .xlsx workbook into CSV-like headers + row dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [str(c).strip() if c is not None else "" for c in header_row]
        # Drop trailing empty header names
        while headers and not headers[-1]:
            headers.pop()
        parsed: list[dict[str, str]] = []
        for values in rows_iter:
            if values is None:
                continue
            row = {}
            empty = True
            for i, header in enumerate(headers):
                if not header:
                    continue
                val = values[i] if i < len(values) else None
                if val is None:
                    row[header] = ""
                else:
                    row[header] = _scalar_str(val)
                    if row[header]:
                        empty = False
            if not empty:
                parsed.append(row)
        return headers, parsed
    finally:
        wb.close()


def xlsx_to_csv_text(data: bytes) -> str:
    headers, rows = parse_xlsx_bytes(data)
    if not headers:
        return ""
    return rows_to_csv_text(headers, rows)


def validate_row(row: dict[str, Any], mapping: dict[str, str | None], row_num: int) -> dict[str, Any]:
    def get(field: str) -> str:
        header = mapping.get(field)
        if not header:
            return ""
        return _scalar_str(row.get(header))

    errors: list[str] = []
    source = _normalize_source(get("source"))
    external_id = get("external_id")
    if not external_id:
        errors.append("Deal ID / external_id obligatorio")
        external_id = f"CSV-{uuid.uuid4().hex[:10].upper()}"
    customer_name = _customer_display_name(get("customer_name"))
    if not customer_name:
        errors.append("customer_name obligatorio")
    try:
        total_value = _dec(get("total_value"))
    except ValueError as exc:
        total_value = Decimal("0")
        errors.append(str(exc))
    try:
        amount_shipping = _dec(get("amount_shipping"))
        qty_dorados = _int(get("qty_dorados"))
        qty_plateados = _int(get("qty_plateados"))
    except Exception as exc:
        amount_shipping = Decimal("0")
        qty_dorados = qty_plateados = 0
        errors.append(str(exc))

    # Filas vacías del export (sin valor): no inventar venta.
    if total_value <= 0:
        errors.append("valor debe ser > 0")
    elif qty_dorados <= 0 and qty_plateados <= 0:
        qty_dorados = 1  # default 1 kit solo si hay valor

    commercial_raw = get("commercial_raw")
    if source == SaleSource.MANUAL and not commercial_raw:
        commercial_raw = "MANUAL"
    if source == SaleSource.FERIAS and not commercial_raw:
        commercial_raw = "FERIAS"
    if source == SaleSource.ECOMMERCE and not commercial_raw:
        commercial_raw = "ECOMMERCE"

    status = (get("status") or "completed").lower()
    if status not in {"processing", "completed", "cancelled", "failed", "refunded", "pending"}:
        status = "completed"

    tracking_number = get("tracking_number")
    tracking_url = get("tracking_url")
    label_url = get("label_url")
    carrier = get("carrier") or "coordinadora"
    shipment_raw = (get("shipment_status") or "").strip().upper()
    # Columna ENVIADO del Excel llega como true/false.
    if shipment_raw in {"TRUE", "1", "SI", "SÍ", "YES"}:
        shipment_status = "ENVIADO"
    elif shipment_raw in {"FALSE", "0", "NO"}:
        shipment_status = "LISTO_PARA_ENVIAR" if tracking_number else "POR_GENERAR"
    else:
        shipment_status = shipment_raw

    # Histórico declarado como ya despachado: con guía siempre ENVIADO (no cola logística).
    if tracking_number:
        shipment_status = "ENVIADO"

    try:
        shipping_cost_raw = get("shipping_cost")
        shipping_cost = _dec(shipping_cost_raw) if shipping_cost_raw else None
    except ValueError as exc:
        shipping_cost = None
        errors.append(str(exc))

    # En este export, TRASPORTE es el costo de guía / flete.
    if shipping_cost is None and amount_shipping > 0 and tracking_number:
        shipping_cost = amount_shipping
    if shipping_cost is not None and amount_shipping == 0:
        amount_shipping = shipping_cost

    # Histórico ya despachado: si hay guía y no dicen estado, asumir ENVIADO.
    if tracking_number and not shipment_status:
        shipment_status = "ENVIADO"
    if shipment_status and shipment_status not in {
        "POR_GENERAR",
        "GUIA_FALLIDA",
        "LISTO_PARA_ENVIAR",
        "ENVIADO",
        "REVISAR",
        "CANCELADA",
    }:
        shipment_status = "ENVIADO"

    # Con guía: forzar ENVIA para que exista Shipment (no regenera Envia).
    fulfillment = normalize_fulfillment_type(
        get("fulfillment_type"),
        requires_shipping=bool(tracking_number or get("address_raw") or get("city_raw")),
    )
    if tracking_number:
        fulfillment = "ENVIA"

    sent_at = None
    if get("sent_at"):
        sent_at = _parse_closed_at(get("sent_at"))
    elif tracking_number and shipment_status == "ENVIADO":
        sent_at = _parse_closed_at(get("closed_at"))

    if tracking_number and not tracking_url:
        tracking_url = f"https://envia.com/es-CO/tracking?label={tracking_number}"

    payload = {
        "external_id": _clip(external_id, 64),
        "source": source,
        "customer_name": _clip(customer_name, 255),
        "email": _clip(get("email"), 254),
        "phone": _sanitize_phone(get("phone")),
        "id_number": _sanitize_id_number(get("id_number")),
        "address_raw": _clip(get("address_raw"), 512),
        "city_raw": _clip(get("city_raw"), 128),
        "state_raw": _clip(get("state_raw"), 128),
        "total_value": total_value,
        "amount_shipping": amount_shipping,
        "payment_account": _clip(get("payment_account"), 128),
        "commercial_raw": _clip(commercial_raw, 128),
        "qty_dorados": qty_dorados,
        "qty_plateados": qty_plateados,
        "tipo_dorados": _clip(normalize_kit_type(get("tipo_dorados")), 128),
        "tipo_plateados": _clip(normalize_kit_type(get("tipo_plateados")), 128),
        "status": status[:64],
        "closed_at": _parse_closed_at(get("closed_at")),
        "order_notes": get("order_notes"),
        "income_source": _clip(source, 32),
        "fulfillment_type": fulfillment,
        "requires_shipping": fulfillment == "ENVIA",
        "tracking_number": _clip(tracking_number, 128),
        "tracking_url": _clip(tracking_url, 512),
        "shipping_cost": shipping_cost,
        "label_url": _clip(label_url, 200),
        "carrier": _clip(carrier, 64),
        "sent_at": sent_at,
        "shipment_status": shipment_status,
    }
    return {"row": row_num, "ok": not errors, "errors": errors, "data": payload}


def dry_run_table(
    headers: list[str],
    rows: list[dict[str, Any]],
    mapping: dict[str, str | None] | None = None,
) -> dict[str, Any]:
    mapping = mapping or detect_mapping(headers)
    results = []
    for i, row in enumerate(rows, start=2):
        results.append(validate_row(row, mapping, i))
    return {
        "headers": headers,
        "mapping": mapping,
        "total": len(results),
        "valid": sum(1 for r in results if r["ok"]),
        "invalid": sum(1 for r in results if not r["ok"]),
        "rows": results,
    }


def dry_run_csv(text: str, mapping: dict[str, str | None] | None = None) -> dict[str, Any]:
    headers, rows = parse_csv_text(text)
    return dry_run_table(headers, rows, mapping=mapping)


def dry_run_xlsx(data: bytes, mapping: dict[str, str | None] | None = None) -> dict[str, Any]:
    headers, rows = parse_xlsx_bytes(data)
    return dry_run_table(headers, rows, mapping=mapping)


def commit_table(
    headers: list[str],
    rows: list[dict[str, Any]],
    *,
    mapping: dict[str, str | None] | None = None,
    on_duplicate: str = "skip",
    actor=None,
) -> dict[str, Any]:
    report = dry_run_table(headers, rows, mapping=mapping)
    created = updated = skipped = rejected = 0
    details: list[dict] = []

    for item in report["rows"]:
        if not item["ok"]:
            rejected += 1
            details.append({"row": item["row"], "status": "rejected", "errors": item["errors"]})
            continue
        data = item["data"]
        source = data["source"]
        model = SOURCE_MODELS[source]
        try:
            with transaction.atomic():
                existing = model.objects.filter(external_id=data["external_id"]).first()
                if existing and on_duplicate == "skip":
                    skipped += 1
                    details.append(
                        {
                            "row": item["row"],
                            "status": "skipped",
                            "external_id": data["external_id"],
                        }
                    )
                    continue

                defaults = {
                    "deal_name": data["customer_name"],
                    "closed_at": data["closed_at"],
                    "total_value": data["total_value"],
                    "amount_shipping": data["amount_shipping"],
                    "payment_account": data["payment_account"],
                    "payment_method": resolve_payment_method(
                        data.get("payment_account") or "", actor=actor
                    ),
                    "income_source": data["income_source"],
                    "status": data["status"],
                    "stage": "Import CSV",
                    "commercial_raw": data["commercial_raw"],
                    "customer_name": data["customer_name"],
                    "email": data["email"],
                    "phone": data["phone"],
                    "id_number": data["id_number"],
                    "address_raw": data["address_raw"],
                    "city_raw": data["city_raw"],
                    "state_raw": data["state_raw"],
                    "qty_dorados": data["qty_dorados"],
                    "qty_plateados": data["qty_plateados"],
                    "tipo_dorados": data.get("tipo_dorados") or "",
                    "tipo_plateados": data.get("tipo_plateados") or "",
                    "order_notes": data["order_notes"],
                    "requires_shipping": data["requires_shipping"],
                    "fulfillment_type": data.get("fulfillment_type") or "ENVIA",
                    "extra": {"imported": True},
                }
                apply_fulfillment(defaults)
                if defaults["payment_method"]:
                    defaults["payment_account"] = defaults["payment_method"].name
                if existing and on_duplicate == "update":
                    for k, v in defaults.items():
                        setattr(existing, k, v)
                    existing.save()
                    source_sale = existing
                    updated += 1
                    status = "updated"
                else:
                    source_sale = model.objects.create(
                        external_id=data["external_id"], **defaults
                    )
                    created += 1
                    status = "created"

                promote_to_consolidated(source_sale, source=source, actor=actor)
                _apply_historical_shipment(
                    source_sale, source=source, data=data, actor=actor
                )
                details.append(
                    {
                        "row": item["row"],
                        "status": status,
                        "external_id": data["external_id"],
                    }
                )
        except Exception as exc:  # noqa: BLE001 — isolate bad historical rows
            rejected += 1
            details.append(
                {
                    "row": item["row"],
                    "status": "rejected",
                    "external_id": data.get("external_id"),
                    "errors": [str(exc)],
                }
            )

    log_audit_event(
        actor=actor,
        action="SALES_CSV_IMPORT",
        entity="ImportJob",
        entity_id="",
        metadata={
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "rejected": rejected,
            "on_duplicate": on_duplicate,
        },
    )
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "rejected": rejected,
        "details": details[:200],
        "mapping": report["mapping"],
    }


def commit_csv(
    text: str,
    *,
    mapping: dict[str, str | None] | None = None,
    on_duplicate: str = "skip",
    actor=None,
) -> dict[str, Any]:
    headers, rows = parse_csv_text(text)
    return commit_table(
        headers, rows, mapping=mapping, on_duplicate=on_duplicate, actor=actor
    )


def commit_xlsx(
    data: bytes,
    *,
    mapping: dict[str, str | None] | None = None,
    on_duplicate: str = "skip",
    actor=None,
) -> dict[str, Any]:
    headers, rows = parse_xlsx_bytes(data)
    return commit_table(
        headers, rows, mapping=mapping, on_duplicate=on_duplicate, actor=actor
    )


def _apply_historical_shipment(source_sale, *, source: str, data: dict, actor=None) -> None:
    """
    Filas con guía ya generada: deja Shipment en ENVIADO con tracking/costo.
    No llama Envia ni descuenta inventario (histórico).
    """
    tracking = (data.get("tracking_number") or "").strip()
    ship_status = (data.get("shipment_status") or "").strip().upper()
    if not tracking and ship_status not in {"ENVIADO", "LISTO_PARA_ENVIAR", "CANCELADA"}:
        return

    from apps.logistics.models import ShipmentStatus
    from apps.logistics.services.shipments import ensure_shipment_for_sale
    from apps.sales.models import ConsolidatedSale

    sale = ConsolidatedSale.objects.filter(source=source, external_id=data["external_id"]).first()
    if not sale:
        return

    # Asegura ENVIA + requires_shipping por si el promote llegó sin guía en signal.
    if tracking and (not sale.requires_shipping or sale.fulfillment_type != "ENVIA"):
        sale.requires_shipping = True
        sale.fulfillment_type = "ENVIA"
        sale.save(update_fields=["requires_shipping", "fulfillment_type", "updated_at"])

    shipment = ensure_shipment_for_sale(sale, actor=actor)
    if not shipment:
        return

    status = ship_status or (ShipmentStatus.ENVIADO if tracking else shipment.status)
    if status not in ShipmentStatus.values:
        status = ShipmentStatus.ENVIADO

    shipment.tracking_number = tracking or shipment.tracking_number
    shipment.tracking_url = (data.get("tracking_url") or "").strip() or shipment.tracking_url
    shipment.label_url = (data.get("label_url") or "").strip() or shipment.label_url
    shipment.carrier = (data.get("carrier") or shipment.carrier or "coordinadora")[:64]
    cost = data.get("shipping_cost")
    if cost is not None:
        shipment.shipping_cost = cost
    elif data.get("amount_shipping"):
        shipment.shipping_cost = data["amount_shipping"]
    shipment.status = status
    if status == ShipmentStatus.ENVIADO:
        shipment.sent_at = data.get("sent_at") or shipment.sent_at or timezone.now()
        shipment.last_error = ""
    # Marca para ocultar de colas operativas de logística/despacho.
    detail = dict(shipment.warning_detail or {})
    detail["historical_import"] = True
    shipment.warning_detail = detail
    shipment.address_mirror = shipment.address_mirror or sale.address_raw
    shipment.city_mirror = shipment.city_mirror or sale.city_raw
    shipment.state_mirror = shipment.state_mirror or sale.state_raw
    shipment.save()

    if shipment.shipping_cost is not None:
        from apps.sales.services.normalization import recalculate_shipping

        recalculate_shipping(sale, shipment.shipping_cost, actor=actor)
